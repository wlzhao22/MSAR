import yaml
import PIL.Image
import numpy as np
import openpyxl
import datetime
import matplotlib.pyplot as plt
import os


def calculate_resized_w_h(w, h):
    if w > h:
        rate = w / 512
        w_resized = 512
        h_resized = int(h / rate)
    else:
        rate = h / 512
        w_resized = int(w / rate)
        h_resized = 512

    return w_resized, h_resized, rate


def get_bboxes_from_file(file_name):
    boxes = []
    file_names = []
    file_boxes = []
    with open(file_name, 'r') as f:
        line = f.readline().strip()
        boxes = []
        while line:
            datas = line.split()
            bbox_name = datas[0].split('>')[0]
            box = list(map(int, datas[1:5]))
            if not file_names:
                file_names.append(bbox_name)
                boxes.append(box)
            else:
                if bbox_name != file_names[-1]:
                    file_names.append(bbox_name)
                    file_boxes.append(boxes)
                    boxes = []
                boxes.append(box)
            line = f.readline().strip()
        file_boxes.append(boxes)

    return file_names, file_boxes

def get_CUHK_SYSU_bboxes_from_file(file_name):
    boxes = []
    file_names = []
    file_boxes = []
    with open(file_name, 'r') as f:
        line = f.readline().strip()
        boxes = []
        while line:
            datas = line.split()
            bbox_name = datas[-1]
            box = list(map(int, datas[1:5]))
            if not file_names:
                file_names.append(bbox_name)
                boxes.append(box)
            else:
                if bbox_name != file_names[-1]:
                    file_names.append(bbox_name)
                    file_boxes.append(boxes)
                    boxes = []
                boxes.append(box)
            line = f.readline().strip()
        file_boxes.append(boxes)

    return file_names, file_boxes

def get_Holidays_bboxes_from_file(file_name):
    file_names = []
    file_boxes = []
    with open(file_name, 'r') as f:
        line = f.readline().strip()
        boxes = []
        while line:
            datas = line.split()
            bbox_name = datas[-1]
            box = list(map(int, datas[1:5]))
            if not file_names:
                file_names.append(bbox_name)
                boxes.append(box)
            else:
                if bbox_name != file_names[-1]:
                    file_names.append(bbox_name)
                    file_boxes.append(boxes)
                    boxes = []
                boxes.append(box)
            line = f.readline().strip()
        file_boxes.append(boxes)

def load_config_file(config_path):
    with open(config_path, 'r') as stream:
        try:
            config = yaml.safe_load(stream)
        except yaml.YAMLError as exc:
            print(exc)
    return config


def load_mask(path):
    fpp = open(path,'r')
    line = fpp.readline().strip()
    mask = []
    while line:
        info = line.split()
        if info[0] == '0':
            mask.append(None)
        else:
            mask.append(info[1])
        line = fpp.readline().strip()
    return mask

def downsample_mask(mask,h,w,min_num=10, rate=1):
    height, width = mask.shape
    # downsampled_height = int(height // (32*rate))
    # downsampled_width = int(width // (32*rate))
    downsampled_height = h
    downsampled_width = w
    patch_h = height // h
    patch_w = width // w

    # 创建一个空的下采样掩码矩阵
    downsampled_mask = np.zeros((downsampled_height, downsampled_width), dtype=int)

    # 遍历每个下采样块
    for i in range(downsampled_height):
        for j in range(downsampled_width):
            # 提取当前块的原始掩码块
            block = mask[i*patch_h:(i+1)*patch_h, j*patch_w:(j+1)*patch_w]
            # print(np.sum(block))

            # 检查当前块是否至少存在一个为1的像素
            if np.sum(block) > 10:
                downsampled_mask[i, j] = 1

    return downsampled_mask

def load_roi_mask(path):
    mask = np.load(path)
    return mask

def save_bi_np_matrix(matrix, f, is_print=False):
    '''
    二进制字节流的形式写入文件
    :param matrix 二维np数组，dim0是数量，dim1是维度:
    :param file_path 文件路径:
    :param is_writen 是否输出，默认不输出:
    :return:
    '''
    # with open(file_path, 'wb') as f:
    for i in range(matrix.shape[0]):
        dim = len(matrix[i])
        np.array(dim, dtype=np.uint32).tofile(f)
        matrix[i].astype(np.float32).tofile(f)
    if is_print:
        print(matrix)

def load_bi_np_matrix(src_file):
    print("Loading Matrix from:", src_file)
    data = np.fromfile(file=src_file, dtype=np.float32)
    data_dim = data.view(np.int32)[0]
    features = data.reshape(-1, data_dim + 1)[:, 1:]
    print("Matrix shape:", features.shape)
    return features

def load_str_np_matrix(src_file):
    features = []
    print("Loading Matrix from:", src_file)
    with open(src_file, 'rb') as f:
        line = f.readline().strip()
        qry_dimension = line
        line = f.readline().strip()
        while line:
            datas = line.split(' ')
            feature = list(map(float, datas))
            features.append(feature)
            line = f.readline().strip()
    features = np.array(features)
    print("Matrix shape:", features.shape)
    return features

def save_str_list(list, file_path, is_print=False):
    '''
    字符串的形式写list
    :param list:
    :param file_path:
    :param is_print:
    :return:
    '''
    if isinstance(file_path, str):
        f = open(file_path, 'w')
    else:
        f = file_path
    for line in list:
        f.write(line)
        if is_print:
            print(line)

def save_str_line(line, f, is_print=False):
    '''
    写入一行文件
    :param list:
    :param file_path:
    :param is_print:
    :return:
    '''
    f.write(line)
    if is_print:
        print(line)

def load_grd_file(file):
    with open(file, 'r') as f:
        line = f.readline().strip()
        grd_dict = {}
        while line:
            data = line.split()
            lt = list(map(str, data[2:]))
            lt = [x for x in lt if x != data[0]]
            grd_dict.update({data[0]: lt})
            line = f.readline().strip()
    f.close()
    return grd_dict

def load_box(file):
    '''
    Load box from qry-box.txt / ref-box.txt
    :param file:
    :return:
    '''
    with open(file, 'r') as f:
        name = []
        cod = []
        line = f.readline().strip()
        while line:
            data = line.split()
            name.append(data[0].split('>')[0])
            cod.append(list(map(int, data[1:])))
            line = f.readline().strip()
    f.close()
    return name, cod

def append_info_to_xlsx(filePath, dataset, info: list):
    assert dataset == "Ins160" or dataset == "Ins335" or dataset == "instre"
    wb = openpyxl.load_workbook(filePath)
    ws = wb[dataset]
    ws.append(info)
    wb.save(filePath)

def append_xlsx(filePath, dataset, info: list):
    assert len(info) == 11
    assert dataset == "Ins160" or dataset == "Ins335" or dataset == "instre"
    wb = openpyxl.load_workbook(filePath)
    ws = wb[dataset]
    ws.append(info)
    wb.save(filePath)

def create_xlsx(filePath):
    wb = openpyxl.Workbook()
    ws = wb['Sheet']
    wb.remove(ws)
    sheetNames = ["Ins160","Ins335","instre"]
    timestamp = datetime.datetime.now()
    for sheetName in sheetNames:
        ws = wb.create_sheet(sheetName)
        row = [timestamp.strftime('%Y-%m-%d')]
        ws.append(row)
        row = ["para", "layer", "method","resized","pooling","whiten","num","dim","map50","map100","mapall"]
        ws.append(row)
    # filePath = root_path + "/" +timestamp.strftime('%Y-%m-%d') + ".xlsx"
    print("Svaing Result to:", filePath)
    wb.save(filePath)

def visualize_img_and_box(img_paths, img_boxes, plt_save=False):
    '''
    可视化多张图片和box
    img_paths 每张图片的绝对路径
    boxes 每张图片对应的框，三维列表 （每张图片对应多个boxes）
    e.g.
    ['/home/ybmiao/data/INSTRE/INSTRE-S1/19a_correction_tape/077.jpg']
    [[[394, 151, 684, 664]]]
    '''
    assert len(img_paths) == len(img_boxes)
    save_path = "./visualization"
    if plt_save:
        if not os.path.exists(save_path):
            os.makedirs(save_path)

    for path, boxes in zip(img_paths, img_boxes):
        img = PIL.Image.open(path, 'r')
        colors = ['r', 'g', 'b', 'y', 'm']
        fig, ax = plt.subplots()
        ax.imshow(img)
        for i in range(len(boxes)):
            ax.add_patch(
                plt.Rectangle((boxes[i][0], boxes[i][1]), boxes[i][2] - boxes[i][0],
                              boxes[i][3] - boxes[i][1], fill=False,
                              edgecolor=colors[i % 5], linewidth=1))
        plt.axis("off")
        if plt_save:
            save_name = "_".join(path.strip().split('.')[0].split('/')[-3:]) + ".png"
            plt.savefig(save_path + "/" + save_name)
        else:
            plt.show()


def load_fvecs(file="", dtype=np.float32, verbose=False):
    """
    this func is designed to read `.fvecs` binary files.
    `.fvecs` :: The vectors are stored in raw little endian.
    Each vector takes 4+d*4 bytes for .fvecs and .ivecs formats, and 4+d bytes for .bvecs formats,
    where d is the dimensionality of the vector, as shown below.
    source link: https://www.gsitechnology.com/ANN-Benchmarks-Data-Scientists-Journey-to-Billion-Scale-Performance
    source link: http://corpus-texmex.irisa.fr/#matlab
    """
    data     = np.fromfile(file=file, dtype=dtype)
    data_dim = data.view(np.int32)[0]
    data     = data.reshape(-1, data_dim + 1)[:, 1:] #FIXME:
    #------------------------------------------------#
    if verbose:
        print('data_name:\t ' + file)
        print('data_dim:\t{}'.format(data_dim))
        print('data_shape:\t {}*{}'.format(data.shape[0], data.shape[1]))
    return data

def load_ivecs(file="", dtype=np.int32, verbose=False):
    data     = np.fromfile(file=file, dtype=dtype)
    data_dim = data.view(np.int32)[0]
    data     = data.reshape(-1, data_dim + 1)[:, 1:]
    #------------------------------------------------#
    if verbose:
        print('data_name:\t ' + file)
        print('data_dim:\t{}'.format(data_dim))
        print('data_shape:\t {}*{}'.format(data.shape[0], data.shape[1]))
    return data


def save_fvecs(data, file):
    size, dim = data.shape
    with open(file, 'wb') as f:
        for i in range(size):
            np.array(dim, dtype=np.uint32).tofile(f)
            data[i].astype(np.float32).tofile(f)

def save_ivecs(data, file):
    size, dim = data.shape
    with open(file, 'wb') as f:
        for i in range(size):
            np.array(dim, dtype=np.uint32).tofile(f)
            data[i].astype(np.uint32).tofile(f)


def save2binary(out_path, data):
    print("Save... {}, its shape: {} X {}".format(out_path, data.shape[0], data.shape[1]))
    with open(out_path, 'wb') as f:
        f.write(data.tobytes(order='C'))
